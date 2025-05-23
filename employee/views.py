from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils import timezone
from .models import Employee, Attendance, Department, Salary, Feedback
import face_recognition
import numpy as np
import cv2
from datetime import date, datetime
import json
from django.contrib.admin.views.decorators import staff_member_required
from django.db import transaction
from django.contrib.auth import authenticate, login, logout
from django.core.paginator import Paginator
from .forms import EmployeeForm
from django.contrib.auth.models import User
from django.db.models import Q, Sum, Avg
import os
from django.http import JsonResponse, HttpResponse
from django.utils.timezone import localtime
from PIL import Image
import io
import logging
import traceback
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from django.urls import reverse

# Set up logging
logger = logging.getLogger(__name__)

def login_view(request):
    if request.user.is_authenticated:
        return redirect('employee:dashboard')
        
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            login(request, user)
            next_url = request.GET.get('next', 'employee:dashboard')
            return redirect(next_url)
        else:
            messages.error(request, 'Tên đăng nhập hoặc mật khẩu không đúng')
    
    return render(request, 'employee/login.html')

def logout_view(request):
    logout(request)
    return redirect('employee:login')

@staff_member_required
def employee_list(request):
    search_query = request.GET.get('search', '')
    department_id = request.GET.get('department', '')
    
    employees = Employee.objects.all()
    
    if search_query:
        employees = employees.filter(
            Q(user__first_name__icontains=search_query) |
            Q(user__last_name__icontains=search_query) |
            Q(employee_id__icontains=search_query)
        )
    
    if department_id:
        employees = employees.filter(department_id=department_id)
    
    # Pagination
    paginator = Paginator(employees, 10)  # Show 10 employees per page
    page_number = request.GET.get('page')
    employees_page = paginator.get_page(page_number)
    
    departments = Department.objects.all()
    
    context = {
        'employees': employees_page,
        'departments': departments,
        'search_query': search_query,
        'selected_department': department_id
    }
    
    return render(request, 'employee/employee_list.html', context)

@login_required
def dashboard(request):
    if request.user.is_staff:
        total_employees = Employee.objects.filter(is_active=True).count()
        today = date.today()
        present_today = Attendance.objects.filter(
            date=today,
            status='present'
        ).count()
        departments = Department.objects.all()
        recent_attendance = Attendance.objects.filter(
            date=today
        ).order_by('-check_in')[:10]
        
        context = {
            'total_employees': total_employees,
            'present_today': present_today,
            'departments': departments,
            'recent_attendance': recent_attendance
        }
        return render(request, 'employee/dashboard.html', context)
    else:
        employee = get_object_or_404(Employee, user=request.user)
        attendance = Attendance.objects.filter(
            employee=employee
        ).order_by('-date')[:10]
        
        # Get today's attendance
        today = timezone.localdate()
        today_attendance = Attendance.objects.filter(
            employee=employee,
            date=today
        ).first()
        
        # Calculate monthly statistics
        first_day = today.replace(day=1)
        if today.month == 12:
            next_month = today.replace(year=today.year + 1, month=1, day=1)
        else:
            next_month = today.replace(month=today.month + 1, day=1)
            
        monthly_attendance = Attendance.objects.filter(
            employee=employee,
            date__gte=first_day,
            date__lt=next_month
        )
        
        total_days = monthly_attendance.filter(status='present').count()
        total_hours = 0
        overtime_hours = 0
        
        for record in monthly_attendance:
            if record.check_in and record.check_out:
                hours = record.calculate_working_hours()
                if hours > employee.standard_work_hours:
                    overtime = hours - employee.standard_work_hours
                    overtime_hours += overtime
                    total_hours += employee.standard_work_hours
                else:
                    total_hours += hours
        
        monthly_stats = {
            'total_days': total_days,
            'total_hours': total_hours,
            'overtime_hours': overtime_hours
        }
        
        return render(request, 'employee/employee_dashboard.html', {
            'employee': employee,
            'attendance': attendance,
            'monthly_stats': monthly_stats,
            'today_attendance': today_attendance
        })

@login_required
def mark_attendance(request):
    if request.method == 'POST' and request.FILES.get('face_image'):
        try:
            # Get the uploaded image
            uploaded_image = face_recognition.load_image_file(request.FILES['face_image'])
            face_locations = face_recognition.face_locations(uploaded_image)
            
            if not face_locations:
                messages.error(request, 'Không phát hiện khuôn mặt trong ảnh')
                return redirect('mark_attendance')
            
            # Get face encoding of the uploaded image
            face_encoding = face_recognition.face_encodings(uploaded_image)[0]
            
            # Get the employee
            employee = get_object_or_404(Employee, user=request.user)
            
            # Get stored face encoding
            stored_encoding = np.frombuffer(employee.face_encoding)
            
            # Compare faces
            matches = face_recognition.compare_faces([stored_encoding], face_encoding)
            face_distances = face_recognition.face_distance([stored_encoding], face_encoding)
            
            if matches[0]:
                confidence = (1 - face_distances[0]) * 100
                
                # Check if attendance already exists for today
                today = date.today()
                attendance, created = Attendance.objects.get_or_create(
                    employee=employee,
                    date=today,
                    defaults={
                        'status': 'present',
                        'check_in': timezone.now(),
                        'face_confidence': confidence
                    }
                )
                
                if not created:
                    if not attendance.check_out:
                        attendance.check_out = timezone.now()
                        attendance.save()
                        messages.success(request, 'Đã ghi nhận giờ ra!')
                    else:
                        messages.info(request, 'Bạn đã chấm công đầy đủ cho hôm nay')
                else:
                    messages.success(request, 'Đã ghi nhận giờ vào!')
            else:
                messages.error(request, 'Xác thực khuôn mặt thất bại')
                
        except Exception as e:
            messages.error(request, f'Lỗi xử lý chấm công: {str(e)}')
        
        return redirect('dashboard')
    
    return render(request, 'employee/mark_attendance.html')

@login_required
def register_face(request):
    if not request.user.is_staff:
        employee = get_object_or_404(Employee, user=request.user)
        
        if request.method == 'POST' and request.FILES.get('face_image'):
            try:
                image = face_recognition.load_image_file(request.FILES['face_image'])
                face_locations = face_recognition.face_locations(image)
                
                if not face_locations:
                    messages.error(request, 'Không phát hiện khuôn mặt trong ảnh')
                    return redirect('register_face')
                
                face_encoding = face_recognition.face_encodings(image)[0]
                employee.face_encoding = face_encoding.tobytes()
                employee.face_image = request.FILES['face_image']
                employee.save()
                
                messages.success(request, 'Đăng ký khuôn mặt thành công!')
                return redirect('dashboard')
                
            except Exception as e:
                messages.error(request, f'Lỗi xử lý ảnh: {str(e)}')
                return redirect('register_face')
        
        return render(request, 'employee/register_face.html', {'employee': employee})
    
    return redirect('dashboard')

@staff_member_required
def salary_list(request):
    # Get filter parameters
    year = request.GET.get('year', datetime.now().year)
    month = request.GET.get('month', datetime.now().month)
    department_id = request.GET.get('department', '')
    
    # Base queryset
    salaries = Salary.objects.filter(year=year, month=month)
    
    # Apply department filter if selected
    if department_id:
        salaries = salaries.filter(employee__department_id=department_id)
    
    # Order by employee name
    salaries = salaries.order_by('employee__user__first_name', 'employee__user__last_name')
    
    # Pagination
    paginator = Paginator(salaries, 10)  # Show 10 salaries per page
    page_number = request.GET.get('page')
    salaries_page = paginator.get_page(page_number)
    
    # Get departments for filter
    departments = Department.objects.all()
    
    context = {
        'salaries': salaries_page,
        'departments': departments,
        'current_year': int(year),
        'current_month': int(month),
        'selected_department': department_id,
        'years': range(2020, datetime.now().year + 1),
        'months': range(1, 13)
    }
    return render(request, 'employee/salary_list.html', context)

@staff_member_required
def generate_salary(request):
    if request.method == 'POST':
        year = int(request.POST.get('year', datetime.now().year))
        month = int(request.POST.get('month', datetime.now().month))
        
        try:
            with transaction.atomic():
                # Get all active employees
                employees = Employee.objects.filter(is_active=True)
                
                for employee in employees:
                    salary_data = employee.calculate_monthly_salary(year, month)
                    
                    # Create or update salary record
                    Salary.objects.update_or_create(
                        employee=employee,
                        year=year,
                        month=month,
                        defaults={
                            'base_pay': salary_data['base_pay'],
                            'regular_hours_pay': salary_data['regular_hours_pay'],
                            'overtime_pay': salary_data['overtime_pay'],
                            'total_salary': salary_data['total_salary'],
                            'total_days': salary_data['total_days'],
                            'total_working_hours': salary_data['total_working_hours'],
                            'overtime_hours': salary_data['overtime_hours']
                        }
                    )
                
                messages.success(request, f'Đã tính lương tháng {month}/{year} cho {employees.count()} nhân viên!')
            
        except Exception as e:
            messages.error(request, f'Lỗi khi tính lương: {str(e)}')
        
        # Redirect back with the same filters
        return redirect(f"{reverse('employee:salary_list')}?month={month}&year={year}")
    
    return redirect('employee:salary_list')

@login_required
def salary_detail(request, salary_id):
    salary = get_object_or_404(Salary, id=salary_id)
    
    # Only allow staff or the employee themselves to view salary details
    if not request.user.is_staff and request.user != salary.employee.user:
        messages.error(request, "Bạn không có quyền xem thông tin lương này.")
        return redirect('employee:dashboard')
    
    # Get attendance records for the salary period
    start_date = date(salary.year, salary.month, 1)
    if salary.month == 12:
        end_date = date(salary.year + 1, 1, 1)
    else:
        end_date = date(salary.year, salary.month + 1, 1)
    
    attendance_records = Attendance.objects.filter(
        employee=salary.employee,
        date__gte=start_date,
        date__lt=end_date
    ).order_by('date')
    
    context = {
        'salary': salary,
        'attendance_records': attendance_records
    }
    return render(request, 'employee/salary_detail.html', context)

@staff_member_required
def add_employee(request):
    if request.method == 'POST':
        form = EmployeeForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                with transaction.atomic():
                    # Create User account with is_staff=False
                    user = User.objects.create_user(
                        username=form.cleaned_data['username'],
                        password=form.cleaned_data['password'],
                        email=form.cleaned_data['email'],
                        first_name=form.cleaned_data['first_name'],
                        last_name=form.cleaned_data['last_name'],
                        is_staff=False  # Set is_staff to False by default
                    )
                    
                    # Create Employee
                    employee = form.save(commit=False)
                    employee.user = user
                    
                    # Handle face image
                    if 'face_image' in request.FILES:
                        employee.face_image = request.FILES['face_image']
                    
                    employee.save()
                    
                    messages.success(request, 'Thêm nhân viên mới thành công!')
                    return redirect('employee:employee_list')
            except Exception as e:
                messages.error(request, f'Lỗi khi tạo nhân viên: {str(e)}')
    else:
        form = EmployeeForm()
    
    return render(request, 'employee/add_employee.html', {'form': form})

@staff_member_required
def edit_employee(request, employee_id):
    employee = get_object_or_404(Employee, id=employee_id)
    user = employee.user

    if request.method == 'POST':
        form = EmployeeForm(request.POST, request.FILES, instance=employee)
        if form.is_valid():
            # Update user information
            user.username = form.cleaned_data['username']
            user.first_name = form.cleaned_data['first_name']
            user.last_name = form.cleaned_data['last_name']
            user.email = form.cleaned_data['email']
            user.is_active = form.cleaned_data['is_active']  # Update user active status
            
            # Update password if provided
            if form.cleaned_data['password']:
                if form.cleaned_data['password'] == form.cleaned_data['confirm_password']:
                    user.set_password(form.cleaned_data['password'])
                else:
                    form.add_error('confirm_password', 'Mật khẩu xác nhận không khớp')
                    return render(request, 'employee/edit_employee.html', {'form': form, 'employee': employee})
            
            user.save()
            employee = form.save()

            # Handle face image update
            if 'face_image' in request.FILES:
                employee.face_image = request.FILES['face_image']
                employee.save()
                # Regenerate face encoding
                try:
                    employee.generate_face_encoding()
                except Exception as e:
                    messages.error(request, f'Không thể tạo dữ liệu khuôn mặt: {str(e)}')

            messages.success(request, 'Cập nhật thông tin nhân viên thành công!')
            return redirect('employee:employee_list')
    else:
        initial_data = {
            'username': user.username,
            'first_name': user.first_name,
            'last_name': user.last_name,
            'email': user.email,
            'is_active': user.is_active,  # Initialize form with current active status
        }
        form = EmployeeForm(instance=employee, initial=initial_data)

    return render(request, 'employee/edit_employee.html', {
        'form': form,
        'employee': employee
    })

@staff_member_required
def manage_attendance(request):
    departments = Department.objects.all()
    selected_department = request.GET.get('department')
    search_query = request.GET.get('search')
    employee_id = request.GET.get('employee_id')
    
    employees = Employee.objects.filter(is_active=True)
    if selected_department:
        employees = employees.filter(department_id=selected_department)
    if search_query:
        employees = employees.filter(
            Q(user__first_name__icontains=search_query) |
            Q(user__last_name__icontains=search_query) |
            Q(employee_id__icontains=search_query)
        )
    
    selected_employee = None
    today_attendance = None
    if employee_id:
        selected_employee = get_object_or_404(Employee, id=employee_id)
        today_attendance = Attendance.objects.filter(
            employee=selected_employee,
            date=date.today()
        ).first()
    
    context = {
        'departments': departments,
        'employees': employees,
        'selected_department': selected_department,
        'search_query': search_query,
        'selected_employee': selected_employee,
        'today_attendance': today_attendance
    }
    
    return render(request, 'employee/manage_attendance.html', context)

@staff_member_required
def admin_mark_attendance(request, employee_id):
    employee = get_object_or_404(Employee, id=employee_id)
    
    if request.method == 'POST':
        status = request.POST.get('status')
        check_in = request.POST.get('check_in')
        check_out = request.POST.get('check_out')
        
        attendance, created = Attendance.objects.get_or_create(
            employee=employee,
            date=date.today(),
            defaults={'status': status}
        )
        
        if check_in:
            attendance.check_in = check_in
        if check_out:
            attendance.check_out = check_out
        
        attendance.status = status
        attendance.save()
        
        messages.success(request, f'Attendance marked for {employee.user.get_full_name()}')
    
    return redirect(f'/employee/manage-attendance/?employee_id={employee_id}')

@staff_member_required
def admin_register_face(request, employee_id):
    employee = get_object_or_404(Employee, id=employee_id)
    
    if request.method == 'POST' and request.FILES.get('face_image'):
        try:
            # Load and process the uploaded image
            image = face_recognition.load_image_file(request.FILES['face_image'])
            face_locations = face_recognition.face_locations(image)
            
            if not face_locations:
                messages.error(request, 'No face detected in the image')
                return redirect(f'/employee/manage-attendance/?employee_id={employee_id}')
            
            # Get face encoding
            face_encoding = face_recognition.face_encodings(image)[0]
            
            # Save the face encoding and image
            employee.face_encoding = face_encoding.tobytes()
            employee.face_image = request.FILES['face_image']
            employee.save()
            
            messages.success(request, f'Face registered for {employee.user.get_full_name()}')
            
        except Exception as e:
            messages.error(request, f'Error processing image: {str(e)}')
    
    return redirect(f'/employee/manage-attendance/?employee_id={employee_id}')

@staff_member_required
def admin_delete_attendance(request, attendance_id):
    attendance = get_object_or_404(Attendance, id=attendance_id)
    employee_id = attendance.employee.id
    attendance.delete()
    messages.success(request, 'Attendance record deleted successfully')
    return redirect(f'/employee/manage-attendance/?employee_id={employee_id}')

@login_required
def auto_mark_attendance(request):
    """View for the automatic face recognition attendance page"""
    context = {}
    if request.user.is_staff:
        # For admin users, provide list of all active employees
        context['employees'] = Employee.objects.filter(is_active=True).order_by('user__first_name', 'user__last_name')
    return render(request, 'employee/auto_mark_attendance.html', context)

def optimize_image(image_file, max_size=(640, 480)):
    """Optimize image size and quality for faster processing"""
    try:
        img = Image.open(image_file)
        
        # Convert to RGB if necessary
        if img.mode != 'RGB':
            img = img.convert('RGB')
        
        # Resize if larger than max_size while maintaining aspect ratio
        if img.size[0] > max_size[0] or img.size[1] > max_size[1]:
            img.thumbnail(max_size, Image.Resampling.LANCZOS)
        
        # Save optimized image to bytes
        img_byte_arr = io.BytesIO()
        img.save(img_byte_arr, format='JPEG', quality=85, optimize=True)
        img_byte_arr.seek(0)
        
        return img_byte_arr, None
    except Exception as e:
        error_msg = f"Image optimization error: {str(e)}"
        logger.error(error_msg)
        logger.error(traceback.format_exc())
        return None, error_msg

# Cache for storing face encodings
face_encoding_cache = {}

def get_face_encoding(employee_id, face_encoding_bytes):
    """Get face encoding from cache or compute and cache it"""
    if employee_id not in face_encoding_cache:
        face_encoding_cache[employee_id] = np.frombuffer(face_encoding_bytes)
    return face_encoding_cache[employee_id]

@login_required
def process_auto_attendance(request):
    """Process the face recognition and mark attendance"""
    if request.method == 'POST' and request.FILES.get('face_image'):
        try:
            logger.info("=== Bắt đầu quá trình nhận diện khuôn mặt ===")
            logger.info(f"Phương thức yêu cầu: {request.method}")
            logger.info(f"File trong yêu cầu: {request.FILES.keys()}")
            logger.info(f"Loại file ảnh: {request.FILES['face_image'].content_type}")
            logger.info(f"Kích thước file: {request.FILES['face_image'].size}")
            
            image_file = request.FILES['face_image']
            if not image_file.content_type.startswith('image/'):
                logger.error(f"Loại file không hợp lệ: {image_file.content_type}")
                return JsonResponse({
                    'success': False,
                    'message': 'Loại file không hợp lệ. Vui lòng tải lên file ảnh.',
                    'error_type': 'invalid_file'
                })

            logger.info("Bước 1: Tối ưu hóa ảnh tải lên")
            optimized_image, error = optimize_image(image_file)
            if error:
                logger.error(f"Lỗi tối ưu hóa ảnh: {error}")
                return JsonResponse({
                    'success': False,
                    'message': error,
                    'error_type': 'optimization_error'
                })
            logger.info("Tối ưu hóa ảnh thành công")
            
            try:
                logger.info("Bước 2: Đang tải ảnh đã tối ưu")
                uploaded_image = face_recognition.load_image_file(optimized_image)
                logger.info(f"Tải ảnh thành công. Kích thước: {uploaded_image.shape}")
            except Exception as e:
                logger.error(f"Lỗi tải ảnh: {str(e)}")
                logger.error(traceback.format_exc())
                return JsonResponse({
                    'success': False,
                    'message': 'Lỗi tải ảnh. Vui lòng thử lại với ảnh khác.',
                    'error_type': 'image_load_error'
                })
            
            try:
                logger.info("Bước 3: Phát hiện khuôn mặt")
                face_locations = face_recognition.face_locations(uploaded_image, model="hog")
                logger.info(f"Số khuôn mặt phát hiện được: {len(face_locations)}")
                
                if not face_locations:
                    logger.warning("Không phát hiện khuôn mặt trong ảnh")
                    return JsonResponse({
                        'success': False,
                        'message': 'Không phát hiện khuôn mặt trong ảnh. Vui lòng đảm bảo khuôn mặt hiển thị rõ ràng.',
                        'error_type': 'no_face_detected'
                    })
                
                if len(face_locations) > 1:
                    logger.warning("Phát hiện nhiều khuôn mặt trong ảnh")
                    return JsonResponse({
                        'success': False,
                        'message': 'Phát hiện nhiều khuôn mặt. Vui lòng chỉ để một khuôn mặt trong khung hình.',
                        'error_type': 'multiple_faces'
                    })
                
                logger.info(f"Vị trí khuôn mặt: {face_locations[0]}")
            except Exception as e:
                logger.error(f"Lỗi phát hiện khuôn mặt: {str(e)}")
                logger.error(traceback.format_exc())
                return JsonResponse({
                    'success': False,
                    'message': 'Lỗi phát hiện khuôn mặt. Vui lòng đảm bảo ánh sáng tốt và khuôn mặt rõ ràng.',
                    'error_type': 'face_detection_error'
                })
            
            try:
                logger.info("Bước 4: Tạo mã hóa khuôn mặt")
                face_encoding = face_recognition.face_encodings(uploaded_image, face_locations)[0]
                logger.info("Tạo mã hóa khuôn mặt thành công")
            except Exception as e:
                logger.error(f"Lỗi mã hóa khuôn mặt: {str(e)}")
                logger.error(traceback.format_exc())
                return JsonResponse({
                    'success': False,
                    'message': 'Lỗi xử lý đặc trưng khuôn mặt. Vui lòng thử lại với ảnh rõ nét hơn.',
                    'error_type': 'encoding_error'
                })

            if not request.user.is_staff:
                current_employee = get_object_or_404(Employee, user=request.user)
                if not current_employee.face_encoding:
                    return JsonResponse({
                        'success': False,
                        'message': 'Chưa đăng ký khuôn mặt cho tài khoản của bạn. Vui lòng liên hệ quản trị viên.',
                        'error_type': 'no_registered_face'
                    })
                
                stored_encoding = np.frombuffer(current_employee.face_encoding)
                matches = face_recognition.compare_faces([stored_encoding], face_encoding)
                face_distances = face_recognition.face_distance([stored_encoding], face_encoding)
                
                if not matches[0]:
                    return JsonResponse({
                        'success': False,
                        'message': 'Khuôn mặt không khớp với khuôn mặt đã đăng ký. Vui lòng đảm bảo bạn đang sử dụng đúng tài khoản của mình.',
                        'error_type': 'face_mismatch'
                    })
                
                confidence = (1 - face_distances[0]) * 100
                if confidence < 60:
                    return JsonResponse({
                        'success': False,
                        'message': 'Độ tin cậy nhận diện khuôn mặt quá thấp. Vui lòng thử lại với điều kiện ánh sáng tốt hơn.',
                        'error_type': 'low_confidence'
                    })
            else:
                employees = Employee.objects.exclude(face_encoding__isnull=True)
                known_face_encodings = []
                known_employees = []
                
                for employee in employees:
                    if employee.face_encoding:
                        known_face_encodings.append(np.frombuffer(employee.face_encoding))
                        known_employees.append(employee)
                
                if not known_face_encodings:
                    return JsonResponse({
                        'success': False,
                        'message': 'Không tìm thấy khuôn mặt đã đăng ký nào trong hệ thống.',
                        'error_type': 'no_registered_faces'
                    })
                
                face_distances = face_recognition.face_distance(known_face_encodings, face_encoding)
                best_match_index = np.argmin(face_distances)
                confidence = (1 - face_distances[best_match_index]) * 100
                
                if confidence >= 60:
                    current_employee = known_employees[best_match_index]
                elif confidence >= 50:
                    current_employee = known_employees[best_match_index]
                    logger.warning(f"Độ tin cậy trung bình ({confidence:.2f}%) cho nhân viên {current_employee.employee_id}")
                else:
                    return JsonResponse({
                        'success': False,
                        'message': f'Độ tin cậy quá thấp ({confidence:.2f}%). Không thể xác định chính xác nhân viên.',
                        'error_type': 'low_confidence'
                    })
            
            today = date.today()
            attendance, created = Attendance.objects.get_or_create(
                employee=current_employee,
                date=today,
                defaults={
                    'status': 'present',
                    'check_in': timezone.now(),
                    'face_confidence': confidence
                }
            )
            
            current_time = localtime()
            
            if not created:
                if not attendance.check_out:
                    attendance.check_out = current_time
                    attendance.save()
                    status = "Đã chấm công ra"
                else:
                    status = "Đã chấm công đủ"
            else:
                status = "Đã chấm công vào"
            
            logger.info(f"Xử lý chấm công thành công cho nhân viên {current_employee.id}")
            
            return JsonResponse({
                'success': True,
                'message': f'Đã nhận diện thành công {current_employee.user.get_full_name()}!',
                'employee_name': current_employee.user.get_full_name(),
                'employee_id': current_employee.employee_id,
                'department': current_employee.department.name,
                'attendance_status': status,
                'timestamp': current_time.strftime('%H:%M'),
                'confidence': f'{confidence:.2f}%'
            })
                
        except Exception as e:
            logger.error(f"Lỗi không mong đợi trong quá trình nhận diện khuôn mặt: {str(e)}")
            logger.error(traceback.format_exc())
            return JsonResponse({
                'success': False,
                'message': f'Lỗi không mong đợi: {str(e)}. Vui lòng thử lại hoặc liên hệ quản trị viên.',
                'error_type': 'unexpected_error'
            })
    
    logger.error("Yêu cầu không hợp lệ: Không có file ảnh")
    return JsonResponse({
        'success': False,
        'message': 'Yêu cầu không hợp lệ. Vui lòng cung cấp ảnh.',
        'error_type': 'invalid_request'
    })

@staff_member_required
def delete_employee(request, employee_id):
    employee = get_object_or_404(Employee, id=employee_id)
    
    try:
        with transaction.atomic():
            # Store user instance before deleting employee
            user = employee.user
            
            # Delete employee record
            employee.delete()
            
            # Delete associated user account
            user.delete()
            
            messages.success(request, f'Đã xóa nhân viên {employee.employee_id} thành công.')
    except Exception as e:
        messages.error(request, f'Lỗi khi xóa nhân viên: {str(e)}')
    
    return redirect('employee:employee_list')

@staff_member_required
def regenerate_face_encoding(request, employee_id):
    """Regenerate face encoding for an employee's existing face image"""
    employee = get_object_or_404(Employee, id=employee_id)
    
    try:
        if employee.face_image:
            # Load the image
            image_path = employee.face_image.path
            image = face_recognition.load_image_file(image_path)
            
            # Detect faces
            face_locations = face_recognition.face_locations(image)
            
            if not face_locations:
                messages.error(request, 'Không phát hiện khuôn mặt trong ảnh đã lưu. Vui lòng tải lên ảnh mới.')
                return redirect('employee:edit_employee', employee_id=employee_id)
            
            # Generate face encoding
            face_encoding = face_recognition.face_encodings(image)[0]
            
            # Save the encoding
            employee.face_encoding = face_encoding.tobytes()
            employee.save()
            
            messages.success(request, f'Tạo lại mã hóa khuôn mặt thành công cho {employee.user.get_full_name()}')
        else:
            messages.error(request, 'Không tìm thấy ảnh khuôn mặt cho nhân viên này')
    except Exception as e:
        messages.error(request, f'Lỗi khi tạo lại mã hóa khuôn mặt: {str(e)}')
    
    return redirect('employee:edit_employee', employee_id=employee_id)

@staff_member_required
def department_list(request):
    """View for displaying all departments and their details"""
    departments = Department.objects.all().order_by('name')
    
    # Get employee count for each department
    for department in departments:
        department.employee_count = Employee.objects.filter(department=department, is_active=True).count()
    
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'add':
            name = request.POST.get('name')
            description = request.POST.get('description')
            if name:
                Department.objects.create(name=name, description=description)
                messages.success(request, 'Phòng ban mới đã được tạo thành công!')
            else:
                messages.error(request, 'Vui lòng nhập tên phòng ban!')
    
    context = {
        'departments': departments,
    }
    return render(request, 'employee/department_list.html', context)

@staff_member_required
def edit_department(request, department_id):
    department = get_object_or_404(Department, id=department_id)
    
    if request.method == 'POST':
        name = request.POST.get('name')
        description = request.POST.get('description')
        
        if name:
            department.name = name
            department.description = description
            department.save()
            messages.success(request, 'Phòng ban đã được cập nhật thành công!')
        else:
            messages.error(request, 'Vui lòng nhập tên phòng ban!')
        
        return redirect('employee:department_list')
    
    context = {
        'department': department
    }
    return render(request, 'employee/edit_department.html', context)

@staff_member_required
def delete_department(request, department_id):
    department = get_object_or_404(Department, id=department_id)
    
    # Check if department has employees
    if Employee.objects.filter(department=department).exists():
        messages.error(request, 'Không thể xóa phòng ban đang có nhân viên!')
    else:
        department.delete()
        messages.success(request, 'Phòng ban đã được xóa thành công!')
    
    return redirect('employee:department_list')

@staff_member_required
def position_list(request):
    """View for managing positions"""
    # Get all unique positions from employees
    positions = Employee.objects.values('position').distinct().order_by('position')
    
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'update':
            old_position = request.POST.get('old_position')
            new_position = request.POST.get('new_position')
            
            if old_position and new_position:
                Employee.objects.filter(position=old_position).update(position=new_position)
                messages.success(request, 'Chức vụ đã được cập nhật thành công!')
            else:
                messages.error(request, 'Vui lòng nhập đầy đủ thông tin!')
    
    context = {
        'positions': positions,
    }
    return render(request, 'employee/position_list.html', context)

@staff_member_required
def attendance_list(request):
    # Get filter parameters
    department_id = request.GET.get('department', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    status = request.GET.get('status', '')
    search_query = request.GET.get('search', '')
    
    # Base queryset
    attendance_records = Attendance.objects.all().order_by('-date', '-check_in')
    
    # Apply filters
    if department_id:
        attendance_records = attendance_records.filter(employee__department_id=department_id)
    
    if date_from:
        try:
            date_from = datetime.strptime(date_from, '%Y-%m-%d').date()
            attendance_records = attendance_records.filter(date__gte=date_from)
        except ValueError:
            messages.error(request, 'Invalid date format for Date From')
    
    if date_to:
        try:
            date_to = datetime.strptime(date_to, '%Y-%m-%d').date()
            attendance_records = attendance_records.filter(date__lte=date_to)
        except ValueError:
            messages.error(request, 'Invalid date format for Date To')
    
    if status:
        attendance_records = attendance_records.filter(status=status)
    
    if search_query:
        attendance_records = attendance_records.filter(
            Q(employee__user__first_name__icontains=search_query) |
            Q(employee__user__last_name__icontains=search_query) |
            Q(employee__employee_id__icontains=search_query)
        )
    
    # Pagination
    paginator = Paginator(attendance_records, 20)  # Show 20 records per page
    page_number = request.GET.get('page')
    attendance_page = paginator.get_page(page_number)
    
    # Get departments for filter
    departments = Department.objects.all()
    
    context = {
        'attendance_records': attendance_page,
        'departments': departments,
        'selected_department': department_id,
        'date_from': date_from if date_from else '',
        'date_to': date_to if date_to else '',
        'selected_status': status,
        'search_query': search_query,
        'status_choices': [
            ('present', 'Có Mặt'),
            ('absent', 'Vắng Mặt'),
            ('late', 'Đi Muộn'),
            ('half_day', 'Nửa Ngày')
        ]
    }
    return render(request, 'employee/attendance_list.html', context)

@staff_member_required
def export_employee_list(request):
    """Export employee list to Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Employee List"

    # Define headers
    headers = ['Employee ID', 'Full Name', 'Department', 'Position', 'Email', 'Phone', 'Join Date', 'Status']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')

    # Add data
    employees = Employee.objects.select_related('user', 'department').all()
    for row, employee in enumerate(employees, 2):
        ws.cell(row=row, column=1, value=employee.employee_id)
        ws.cell(row=row, column=2, value=employee.user.get_full_name())
        ws.cell(row=row, column=3, value=employee.department.name if employee.department else '')
        ws.cell(row=row, column=4, value=employee.position)
        ws.cell(row=row, column=5, value=employee.user.email)
        ws.cell(row=row, column=6, value=employee.phone_number)
        ws.cell(row=row, column=7, value=employee.joining_date.strftime('%Y-%m-%d') if employee.joining_date else '')
        ws.cell(row=row, column=8, value='Active' if employee.is_active else 'Inactive')

    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width

    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=employee_list.xlsx'
    wb.save(response)
    return response

@staff_member_required
def export_attendance_list(request):
    """Export attendance list to Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance List"

    # Define headers
    headers = ['Date', 'Employee ID', 'Employee Name', 'Department', 'Check In', 'Check Out', 'Status', 'Working Hours']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')

    # Get filter parameters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    department_id = request.GET.get('department')
    
    # Query attendance records
    attendance_records = Attendance.objects.select_related('employee', 'employee__user', 'employee__department').all()
    
    if start_date:
        attendance_records = attendance_records.filter(date__gte=start_date)
    if end_date:
        attendance_records = attendance_records.filter(date__lte=end_date)
    if department_id:
        attendance_records = attendance_records.filter(employee__department_id=department_id)

    # Add data
    for row, record in enumerate(attendance_records, 2):
        ws.cell(row=row, column=1, value=record.date.strftime('%Y-%m-%d'))
        ws.cell(row=row, column=2, value=record.employee.employee_id)
        ws.cell(row=row, column=3, value=record.employee.user.get_full_name())
        ws.cell(row=row, column=4, value=record.employee.department.name if record.employee.department else '')
        ws.cell(row=row, column=5, value=record.check_in.strftime('%H:%M:%S') if record.check_in else '')
        ws.cell(row=row, column=6, value=record.check_out.strftime('%H:%M:%S') if record.check_out else '')
        ws.cell(row=row, column=7, value=record.status.title())
        ws.cell(row=row, column=8, value=round(record.calculate_working_hours(), 2) if record.check_in and record.check_out else 0)

    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width

    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=attendance_list.xlsx'
    wb.save(response)
    return response

@staff_member_required
def export_salary_list(request):
    """Export salary list to Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Salary List"

    # Define headers
    headers = ['Month', 'Employee ID', 'Employee Name', 'Department', 'Base Pay', 'Regular Hours Pay', 
              'Overtime Pay', 'Total Salary', 'Total Days', 'Working Hours', 'Overtime Hours']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')

    # Get filter parameters
    month = request.GET.get('month')
    year = request.GET.get('year')
    department_id = request.GET.get('department')
    
    # Query salary records
    salary_records = Salary.objects.select_related('employee', 'employee__user', 'employee__department').all()
    
    if month and year:
        salary_records = salary_records.filter(month=month, year=year)
    if department_id:
        salary_records = salary_records.filter(employee__department_id=department_id)

    # Add data
    for row, record in enumerate(salary_records, 2):
        ws.cell(row=row, column=1, value=f"{record.year}-{record.month:02d}")
        ws.cell(row=row, column=2, value=record.employee.employee_id)
        ws.cell(row=row, column=3, value=record.employee.user.get_full_name())
        ws.cell(row=row, column=4, value=record.employee.department.name if record.employee.department else '')
        ws.cell(row=row, column=5, value=float(record.base_pay))
        ws.cell(row=row, column=6, value=float(record.regular_hours_pay))
        ws.cell(row=row, column=7, value=float(record.overtime_pay))
        ws.cell(row=row, column=8, value=float(record.total_salary))
        ws.cell(row=row, column=9, value=record.total_days)
        ws.cell(row=row, column=10, value=float(record.total_working_hours))
        ws.cell(row=row, column=11, value=float(record.overtime_hours))

    # Add summary row
    summary_row = ws.max_row + 2
    ws.cell(row=summary_row, column=1, value="Total").font = Font(bold=True)
    for col in range(5, 9):  # Columns E to H (5 to 8) - monetary values
        column_letter = ws.cell(row=1, column=col).column_letter
        ws.cell(row=summary_row, column=col, value=f"=SUM({column_letter}2:{column_letter}{ws.max_row-2})").font = Font(bold=True)

    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width

    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=salary_list.xlsx'
    wb.save(response)
    return response

@login_required
def check_in(request):
    if request.method == 'POST':
        employee = get_object_or_404(Employee, user=request.user)
        today = timezone.localdate()
        
        attendance, created = Attendance.objects.get_or_create(
            employee=employee,
            date=today,
            defaults={
                'status': 'present',
                'check_in': timezone.now()
            }
        )
        
        if created:
            messages.success(request, 'Chấm công vào thành công!')
        else:
            if attendance.check_in:
                messages.info(request, 'Bạn đã chấm công vào hôm nay rồi!')
            else:
                attendance.check_in = timezone.now()
                attendance.status = 'present'
                attendance.save()
                messages.success(request, 'Chấm công vào thành công!')
                
    return redirect('employee:dashboard')

@login_required
def check_out(request):
    if request.method == 'POST':
        employee = get_object_or_404(Employee, user=request.user)
        today = timezone.localdate()
        
        try:
            attendance = Attendance.objects.get(
                employee=employee,
                date=today
            )
            
            if attendance.check_out:
                messages.info(request, 'Bạn đã chấm công ra hôm nay rồi!')
            elif not attendance.check_in:
                messages.error(request, 'Bạn chưa chấm công vào!')
            else:
                attendance.check_out = timezone.now()
                attendance.save()
                messages.success(request, 'Chấm công ra thành công!')
        except Attendance.DoesNotExist:
            messages.error(request, 'Không tìm thấy bản ghi chấm công cho hôm nay!')
            
    return redirect('employee:dashboard')

@login_required
def submit_feedback(request):
    if request.method == 'POST':
        employee = get_object_or_404(Employee, user=request.user)
        feedback_type = request.POST.get('feedback_type')
        content = request.POST.get('content')
        
        if feedback_type and content:
            Feedback.objects.create(
                employee=employee,
                feedback_type=feedback_type,
                content=content,
                submitted_at=timezone.now()
            )
            messages.success(request, 'Cảm ơn bạn đã gửi phản hồi!')
        else:
            messages.error(request, 'Vui lòng điền đầy đủ thông tin!')
            
    return redirect('employee:dashboard')

@staff_member_required
def feedback_list(request):
    # Get filter parameters
    department_id = request.GET.get('department', '')
    feedback_type = request.GET.get('feedback_type', '')
    status = request.GET.get('status', '')
    search_query = request.GET.get('search', '')
    
    # Base queryset
    feedbacks = Feedback.objects.select_related('employee', 'employee__user', 'employee__department').all()
    
    # Apply filters
    if department_id:
        feedbacks = feedbacks.filter(employee__department_id=department_id)
    
    if feedback_type:
        feedbacks = feedbacks.filter(feedback_type=feedback_type)
    
    if status:
        if status == 'resolved':
            feedbacks = feedbacks.filter(is_resolved=True)
        elif status == 'unresolved':
            feedbacks = feedbacks.filter(is_resolved=False)
    
    if search_query:
        feedbacks = feedbacks.filter(
            Q(employee__user__first_name__icontains=search_query) |
            Q(employee__user__last_name__icontains=search_query) |
            Q(employee__employee_id__icontains=search_query)
        )
    
    # Order by submission time (newest first)
    feedbacks = feedbacks.order_by('-submitted_at')
    
    # Pagination
    paginator = Paginator(feedbacks, 10)  # Show 10 feedbacks per page
    page_number = request.GET.get('page')
    feedbacks_page = paginator.get_page(page_number)
    
    # Get departments for filter
    departments = Department.objects.all()
    
    context = {
        'feedbacks': feedbacks_page,
        'departments': departments,
        'selected_department': department_id,
        'selected_type': feedback_type,
        'selected_status': status,
        'search_query': search_query,
        'feedback_types': Feedback.FEEDBACK_TYPES
    }
    
    return render(request, 'employee/feedback_list.html', context)

@staff_member_required
def resolve_feedback(request, feedback_id):
    feedback = get_object_or_404(Feedback, id=feedback_id)
    
    if request.method == 'POST':
        resolution_notes = request.POST.get('resolution_notes')
        if resolution_notes:
            feedback.resolve(notes=resolution_notes)
            messages.success(request, 'Phản hồi đã được xử lý thành công!')
            # Check if request is AJAX
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'status': 'success', 'message': 'Phản hồi đã được xử lý thành công!'})
        else:
            messages.error(request, 'Vui lòng nhập ghi chú xử lý!')
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'status': 'error', 'message': 'Vui lòng nhập ghi chú xử lý!'}, status=400)
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'status': 'error', 'message': 'Phương thức không được hỗ trợ!'}, status=405)
    return redirect('employee:feedback_detail', feedback_id=feedback_id)

@staff_member_required
def export_feedback_list(request):
    """Export feedback list to Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Danh Sách Phản Hồi"

    # Define headers
    headers = [
        'Thời Gian', 'Mã NV', 'Họ Tên', 'Phòng Ban', 
        'Loại Phản Hồi', 'Nội Dung', 'Trạng Thái', 
        'Thời Gian Xử Lý', 'Ghi Chú Xử Lý'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')

    # Get filter parameters
    department_id = request.GET.get('department')
    feedback_type = request.GET.get('feedback_type')
    status = request.GET.get('status')
    search_query = request.GET.get('search')
    
    # Query feedback records
    feedbacks = Feedback.objects.select_related('employee', 'employee__user', 'employee__department').all()
    
    if department_id:
        feedbacks = feedbacks.filter(employee__department_id=department_id)
    if feedback_type:
        feedbacks = feedbacks.filter(feedback_type=feedback_type)
    if status:
        if status == 'resolved':
            feedbacks = feedbacks.filter(is_resolved=True)
        elif status == 'unresolved':
            feedbacks = feedbacks.filter(is_resolved=False)
    if search_query:
        feedbacks = feedbacks.filter(
            Q(employee__user__first_name__icontains=search_query) |
            Q(employee__user__last_name__icontains=search_query) |
            Q(employee__employee_id__icontains=search_query)
        )

    # Add data
    for row, feedback in enumerate(feedbacks, 2):
        ws.cell(row=row, column=1, value=feedback.submitted_at.strftime('%d/%m/%Y %H:%M'))
        ws.cell(row=row, column=2, value=feedback.employee.employee_id)
        ws.cell(row=row, column=3, value=feedback.employee.user.get_full_name())
        ws.cell(row=row, column=4, value=feedback.employee.department.name)
        ws.cell(row=row, column=5, value=feedback.get_feedback_type_display())
        ws.cell(row=row, column=6, value=feedback.content)
        ws.cell(row=row, column=7, value='Đã Xử Lý' if feedback.is_resolved else 'Chưa Xử Lý')
        ws.cell(row=row, column=8, value=feedback.resolved_at.strftime('%d/%m/%Y %H:%M') if feedback.resolved_at else '')
        ws.cell(row=row, column=9, value=feedback.resolution_notes if feedback.resolution_notes else '')

    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width

    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=danh_sach_phan_hoi.xlsx'
    wb.save(response)
    return response

@staff_member_required
def feedback_detail(request, feedback_id):
    feedback = get_object_or_404(Feedback, id=feedback_id)
    return render(request, 'employee/feedback_detail.html', {
        'feedback': feedback
    })
